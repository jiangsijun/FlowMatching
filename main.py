# encoding: utf-8
import argparse
import sys
# sys.path.append('FlowHelper')
# sys.path.append('DispHelper')
sys.path.append('DepthHelper')
# sys.path.append('FileHelper')

import torch
import openpyxl


from FileHelper.FileLoadHelper import SynDatasetLoader
from DepthHelper import MetricHelper
from DepthHelper.MetricHelper import MetricHelper
from DepthHelper.params.SystemParams import SystemParam
from DepthHelper.params.Intrinsics import IntrinsicsParam
from DepthHelper.params.Extrinsics import ExtrinsicsParam
# from UniflowHelper.UniFlowHelper import UniFlowHelper
# from UniflowHelper.UniDispHelper import UniDispHelper
# from DLNRHelper.DispHelper import  DLNRHelper
# from MochaStereo.MochaHelper import MochaHelper
from CrocoHelper.CrocoFlowHelper import CrocoFlowHelper
from CrocoHelper.CrocoDispHelper import CrocoDispHelper


PI = 3.1415926

# 保存并显示输出
class Tee:
    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)

    def flush(self):
        for f in self.files:
            f.flush()

def main(args):

    with open('output.txt', 'w') as f:
        original_stdout = sys.stdout
        sys.stdout = Tee(sys.stdout, f)

        loader = SynDatasetLoader(args.input_dir)
        dispResult = torch.zeros((4, loader.SumLen ))
        flowResult = torch.zeros((4, loader.SumLen ))

        evaluate = MetricHelper(1920, 1280, None)

        DEVICE = 'cuda' #if torch.cuda.is_available() else 'cpu'
        disp = CrocoDispHelper(args.disp_model_path, args, DEVICE)

        flow = CrocoFlowHelper(args.flow_model_path, args, DEVICE)

        for i in range(loader.SumLen):
            depth, left, right, rightF, rightMV, rightRT, hardware = loader.LoadData()

            print("--------------------------part", i, ":", loader.sourceList[i]["rightMV"])
            # print("right")
            sysParam = inputSystemParam(hardware[0], hardware[1], DEVICE)
            evaluate.SetSysparam(sysParam)
            disp_up = disp.Estim(left, right, args.valid_iters)
            flow_up = flow.Estim(left, right, args.valid_iters)
            tdisp, tflow , dispColor , flowColor = evaluate.Compare(disp_up, flow_up, depth)
            # cv2.imwrite("demo_output/dispRight.png" , dispColor)
            # cv2.imwrite("demo_output/flowRight.png" , flowColor)
            dispResult[0][i] = tdisp
            flowResult[0][i] = tflow

            # print("rightF")
            sysParam = inputSystemParam(hardware[0], hardware[1], DEVICE)
            evaluate.SetSysparam(sysParam)
            disp_up = disp.Estim(left, rightF, args.valid_iters)
            flow_up = flow.Estim(left, rightF, args.valid_iters)
            tdisp, tflow , dispColor , flowColor= evaluate.Compare(disp_up, flow_up, depth)
            # cv2.imwrite("demo_output/dispRightF.png" , dispColor)
            # cv2.imwrite("demo_output/flowRightF.png" , flowColor)
            dispResult[1][i] = tdisp
            flowResult[1][i] = tflow

            # print("rightMV")
            sysParam = inputSystemParam(hardware[0], hardware[1], DEVICE)
            evaluate.SetSysparam(sysParam)
            disp_up = disp.Estim(left, rightMV, args.valid_iters)
            flow_up = flow.Estim(left, rightMV, args.valid_iters)
            tdisp, tflow , dispColor , flowColor= evaluate.Compare(disp_up, flow_up, depth)
            # cv2.imwrite("demo_output/dispRightMV.png" , dispColor)
            # cv2.imwrite("demo_output/flowRightMV.png" , flowColor)
            dispResult[2][i] = tdisp
            flowResult[2][i] = tflow

            # print("rightRT")
            sysParam = inputSystemParam(hardware[0], hardware[1], DEVICE)
            evaluate.SetSysparam(sysParam)
            disp_up = disp.Estim(left, rightRT, args.valid_iters)
            flow_up = flow.Estim(left, rightRT, args.valid_iters)
            tdisp, tflow , dispColor , flowColor= evaluate.Compare(disp_up, flow_up, depth)
            # cv2.imwrite("demo_output/dispRightRT.png" , dispColor)
            # cv2.imwrite("demo_output/flowRightRT.png" , flowColor)
            dispResult[3][i] = tdisp
            flowResult[3][i] = tflow

            sys.stdout = original_stdout

    print("disp avg:" , torch.mean(dispResult,dim=1),"flow avg:" , torch.mean(flowResult,dim=1))

    dispVar = torch.zeros((3,loader.SumLen))
    flowVar = torch.zeros((3,loader.SumLen))
    dispVar[0] = (dispResult[1]-dispResult[0])**2
    dispVar[1] = (dispResult[2]-dispResult[0])**2
    dispVar[2] = (dispResult[3] - dispResult[0]) ** 2
    flowVar[0] = (flowResult[1]-flowResult[0])**2
    flowVar[1] = (flowResult[2] - flowResult[0]) ** 2
    flowVar[2] = (flowResult[3] - flowResult[0]) ** 2
    print("disp Var:", torch.mean(dispVar, dim=1), "flow Var:", torch.mean(flowVar, dim=1))

    dispVar[0] = torch.abs(dispResult[1]-dispResult[0])
    dispVar[1] = torch.abs(dispResult[2]-dispResult[0])
    dispVar[2] = torch.abs(dispResult[3] - dispResult[0])
    flowVar[0] = torch.abs(flowResult[1]-flowResult[0])
    flowVar[1] = torch.abs(flowResult[2] - flowResult[0])
    flowVar[2] = torch.abs(flowResult[3] - flowResult[0])

    print("disp abs:" ,torch.mean(dispVar,dim=1) , "flow abs:" ,torch.mean(flowVar,dim=1))

    op_toExcel(dispResult , flowResult , "crocoresult_95_1_0.xlsx")
    #  Var




    return

def inputSystemParam(hardware1, hardware2, device):
    leftCameraIntrins = IntrinsicsParam(hardware1[2][0], hardware1[2][0], hardware1[2][1], hardware1[2][2])
    rightCameraIntrins = IntrinsicsParam(hardware2[2][0], hardware2[2][0], hardware2[2][1], hardware2[2][2])
    diff = hardware2 - hardware1
    tMatrix = torch.tensor([diff[0]], device=device).transpose(1, 0)
    coscita = torch.cos(torch.tensor(diff[1] / 180 * PI, device=device))
    sincita = torch.sin(torch.tensor(diff[1] / 180 * PI, device=device))
    Rx = torch.tensor([[1, 0, 0], [0, coscita[0], -sincita[0]], [0, sincita[0], coscita[0]]], device=device)
    Ry = torch.tensor([[coscita[1], 0, sincita[1]], [0, 1, 0], [-sincita[1], 0, coscita[1]]], device=device)
    Rz = torch.tensor([[coscita[2], -sincita[2], 0], [sincita[2], coscita[2], 0], [0, 0, 1]], device=device)
    RMatrix = Rz @ Ry @ Rx
    cameraExt = ExtrinsicsParam(RMatrix, tMatrix)

    return SystemParam(None, cameraExt, leftCameraIntrins, rightCameraIntrins, None)

def op_toExcel(dispData , flowData , fileName):  # openpyxl库储存数据到excel
    wb = openpyxl.Workbook()  # 创建工作簿对象
    wb.create_sheet("disp",1)
    wb.create_sheet("flow",2)
    wdisp = wb[('disp')]  # 创建子表
    wflow = wb[('flow')]
    wdisp.append(["No." , "Right" , "RightF" , "RightMV" , "RightRT"])
    wflow.append(["No." , "Right" , "RightF" , "RightMV" , "RightRT"])

    for i in range(len(dispData[0])):
        wdisp.append([str(i),str(dispData[0][i].item()),str(dispData[1][i].item()),str(dispData[2][i].item()),str(dispData[3][i].item())])
        wflow.append([str(i),str(flowData[0][i].item()),str(flowData[1][i].item()),str(flowData[2][i].item()),str(flowData[3][i].item())])


    wb.save(fileName)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('--hidden_dims', nargs='+', type=int, default=[128] * 3, help="hidden state and context dimensions")
    parser.add_argument('--corr_implementation', choices=["reg", "alt", "reg_cuda", "alt_cuda"], default="reg", help="correlation volume implementation")
    parser.add_argument('--shared_backbone', action='store_true', help="use a single backbone for the context and feature encoders")
    parser.add_argument('--corr_levels', type=int, default=2, help="number of levels in the correlation pyramid") # Mocha:2 ,Other:4
    parser.add_argument('--corr_radius', type=int, default=4, help="width of the correlation pyramid")
    parser.add_argument('--n_downsample', type=int, default=2, help="resolution of the disparity field (1/2^K)")
    parser.add_argument('--slow_fast_gru', action='store_true', help="iterate the low-res GRUs more frequently")
    parser.add_argument('--n_gru_layers', type=int, default=3, help="number of hidden GRU levels")
    parser.add_argument('--valid_iters', type=int, default=32, help='number of flow-field updates during forward pass')
    parser.add_argument('--mixed_precision', action='store_true', help='use mixed precision')
    parser.add_argument('--input_dir', type=str, default="./dataset")
    parser.add_argument('--flow_model_path', default='./raft/crocoflow.pth', help='location of depth estim model')
    parser.add_argument('--disp_model_path', default='./raft/crocostereo.pth', help='location of disp estim model')
    parser.add_argument('--small', default=False)
    # Mocha
    parser.add_argument('--max_disp', type=int, default=192, help="max disp of geometry encoding volume") #mocha
    parser.add_argument('--tile_overlap', type=float, default=0.9, help='overlap between tiles')   # croco

    args = parser.parse_args()
    main(args)